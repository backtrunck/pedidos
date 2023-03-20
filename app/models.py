import os

from io import BytesIO
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from tempfile import NamedTemporaryFile

from app import db


class Categoria(db.Model):
    __tablename__ = 'categoria'
    id = db.Column(db.Integer, primary_key=True)
    ds_categoria = db.Column(db.String(200))\


    @classmethod
    def buscar(cls, parametros):
        query = cls.query.order_by(cls.ds_categoria)
        for field_key, field_value in parametros.items():
            if field_key == 'id':
                query = query.filter(cls.id == field_value)
            elif field_key == 'ds_categoria':
                query = query.filter(cls.ds_categoria.like(f'%{field_value}%'))
        return query

    @classmethod
    def gerar_xlsx(cls, parametros, query=None):
        if query is None:
            # se a query veio vazia traz todos os dados
            query = cls.buscar(parametros)
        dados = query.all()
        if dados:
            wb = Workbook()
            ws1 = wb.active
            # cabeçalho do arquivo excel
            ws1.append(
                ['id', 'ds_categoria'])
            for dado in dados:
                # linhas do arquivo excel
                ds_categoria = ILLEGAL_CHARACTERS_RE.sub(r'', dado.ds_categoria)
                ws1.append([dado.id, dado.ds_categoria])

            with NamedTemporaryFile(delete=False) as tmp:
                wb.save(tmp.name)
                output = BytesIO(tmp.read())
                tmp.close()
                os.unlink(tmp.name)
            return output


class Cliente(db.Model):
    __tablename__ = 'cliente'
    id = db.Column(db.Integer, primary_key=True)
    cpf = db.Column(db.String(11))
    nm_cliente = db.Column(db.String(200))

    @classmethod
    def buscar(cls, parametros):
        query = cls.query.order_by(cls.nm_cliente)
        for field_key, field_value in parametros.items():
            if field_key == 'id':
                query = query.filter(cls.id == field_value)
            elif field_key == 'cpf':
                query = query.filter(cls.cpf.like(f'%{field_value}%'))
            elif field_key == 'nm_cliente':
                query = query.filter(cls.nm_cliente.like(f'%{field_value}%'))

        return query

    @classmethod
    def gerar_xlsx(cls, parametros, query=None):
        if query is None:
            # se a query veio vazia traz todos os dados
            query = cls.buscar(parametros)
        dados = query.all()
        if dados:
            wb = Workbook()
            ws1 = wb.active
            # cabeçalho do arquivo excel
            ws1.append(
                ['id', 'cpf', 'nm_cliente'])
            for dado in dados:
                # linhas do arquivo excel
                nm_cliente = ILLEGAL_CHARACTERS_RE.sub(r'', dado.nm_cliente)
                ws1.append([dado.id, dado.cpf, dado.nm_cliente])

            with NamedTemporaryFile(delete=False) as tmp:
                wb.save(tmp.name)
                output = BytesIO(tmp.read())
                tmp.close()
                os.unlink(tmp.name)
            return output


class Produto(db.Model):
    __tablename__ = 'produto'
    id = db.Column(db.Integer, primary_key=True)
    cod_barras = db.Column(db.String(15))
    ds_produto = db.Column(db.String(200))
    categoria_id = db.Column(db.Integer)

    @classmethod
    def buscar(cls, parametros):
        query = cls.query.order_by(cls.ds_produto)
        for field_key, field_value in parametros.items():
            if field_key == 'id':
                query = query.filter(cls.id == field_value)
            elif field_key == 'cod_barras':
                query = query.filter(cls.cod_barras.like(f'%{field_value}%'))
            elif field_key == 'ds_produto':
                query = query.filter(cls.ds_produto.like(f'%{field_value}%'))
            elif field_key == 'categoria_id':
                query = query.filter(cls.categoria_id == field_value)

        return query

    @classmethod
    def gerar_xlsx(cls, parametros, query=None):
        if query is None:
            # se a query veio vazia traz todos os dados
            query = cls.buscar(parametros)
        dados = query.all()
        if dados:
            wb = Workbook()
            ws1 = wb.active
            # cabeçalho do arquivo excel
            ws1.append(
                ['id', 'código_barras', 'descrição', 'categoria_id'])
            for dado in dados:
                # linhas do arquivo excel
                ds_produto = ILLEGAL_CHARACTERS_RE.sub(r'', dado.ds_produto)
                ws1.append([dado.id, dado.cod_barras, dado.ds_produto, dado.categoria_id])

            with NamedTemporaryFile(delete=False) as tmp:
                wb.save(tmp.name)
                output = BytesIO(tmp.read())
                tmp.close()
                os.unlink(tmp.name)
            return output
